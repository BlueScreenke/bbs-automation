"""
export/excel_exporter.py

Step 6E — writes the final Bar Bending Schedule to Excel.

Layout and formula conventions deliberately mirror the project's own
reference template (Thome_Beam_BBS_Typical_floors.xlsm), per direct
instruction:

  - Main "BBS" sheet: one row per bar occurrence (marks are never merged
    — see bar_matcher.py's own "multi-occurrence marks" decision, kept
    consistent all the way through to this final output), columns
    Member .. Total Length. "Length of each bar" and "Total Length" are
    Excel FORMULAS (SUM of A:E, and that times Total No.), never a
    Python-computed literal — the sheet recalculates if any input cell
    changes, exactly as the reference template does it.

  - A separate "Hook Parameters" sheet (Shape | Steel | A (mm) | Key) is
    the ONLY place a hook length lives. The main sheet's A/C columns for
    shape 11/21 bars read this table via INDEX/MATCH, using the
    "{shape}_T{diameter}" key shape_resolver.py already generated per
    bar. Changing a hook length in the future means editing this sheet
    directly — never the Python code (explicit project decision).

  - A "Steel Summary" sheet totals length and weight by bar diameter,
    via SUMIF over the main sheet's own Total Length column and the
    standard BS4449 unit-weight relationship (d^2/162 kg/m) — a physical
    constant, not a project assumption, so it's written directly into
    the formula rather than added to the editable Hook Parameters table.

This module computes nothing about a bar's own length or shape — every
number and lookup key it writes came from length_calculator.py and
shape_resolver.py (via Bar, from beam_converter.py). Its only job is
laying that out on a sheet, as formulas, in the reference template's own
convention.

Public surface
--------------
    export_to_excel(beams, output_path) -> None
"""

from __future__ import annotations

from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from models.beam import Beam

FONT_NAME = "Arial"

HEADER_FONT   = Font(name=FONT_NAME, bold=True)
TITLE_FONT    = Font(name=FONT_NAME, bold=True, size=14)
NOTE_FONT     = Font(name=FONT_NAME, italic=True, size=9)
INPUT_FONT    = Font(name=FONT_NAME, color="0000FF")   # blue — editable input, per project convention
INPUT_FILL    = PatternFill("solid", fgColor="FFFF00")  # yellow — cell the reader should fill in / can edit
HEADER_FILL   = PatternFill("solid", fgColor="D9D9D9")
CENTER        = Alignment(horizontal="center", vertical="center", wrap_text=True)
BODY_FONT     = Font(name=FONT_NAME)

BBS_HEADERS = [
    "Member", "Bar mark", "Type and size", "No. of members",
    "No. of bars in each member", "Total No.", "Length of each bar (mm)",
    "Shape code", "A", "B", "C", "D", "E", "Total Length (mm)",
]

# Seed values for the editable hook-parameter table — confirmed against
# the reference workbook's own table (T16/T20 -> 250mm, T25 -> 300mm)
# and extended to cover shape 21 (the reference only had rows for shape
# 11 and typed 21's identical values in by hand — see shape_resolver.py's
# own docstring). These are a starting point only: the whole point of
# this sheet, per project decision, is that they are edited here in
# Excel from now on, never in Python.
HOOK_PARAMETER_SEED = [
    ("11", "T16", 250),
    ("11", "T20", 250),
    ("11", "T25", 300),
    ("21", "T16", 250),
    ("21", "T20", 250),
    ("21", "T25", 300),
]

STEEL_UNIT_WEIGHT_DIVISOR = 162   # BS4449 standard: kg/m = d^2/162 — a physical constant, not project data


# ── Public API ────────────────────────────────────────────────────────────────

def export_to_excel(beams: list[Beam], output_path: str) -> None:
    """
    Write the full Bar Bending Schedule to an .xlsx file.

    Args:
        beams:        Output of beam_converter.convert_to_beams() — every
                       Bar already carries its resolved shape code and
                       A/B/C/D dimensions (values and/or hook lookup keys).
        output_path:  Path to write the .xlsx file to.
    """
    wb = Workbook()

    hook_ws = wb.active
    hook_ws.title = "Hook Parameters"
    hook_range_a, hook_range_key = _write_hook_parameters_sheet(hook_ws)

    bbs_ws = wb.create_sheet("BBS")
    last_row = _write_bbs_sheet(bbs_ws, beams, hook_range_a, hook_range_key)

    summary_ws = wb.create_sheet("Steel Summary")
    _write_steel_summary_sheet(summary_ws, beams, last_row)

    for ws in (hook_ws, bbs_ws, summary_ws):
        _apply_default_font(ws)

    wb.save(output_path)


def _apply_default_font(ws: Worksheet) -> None:
    """Ensures every cell that wasn't explicitly styled still renders in
    the project's professional font (Arial), not openpyxl's default."""
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None and cell.font.name != FONT_NAME:
                cell.font = BODY_FONT


# ── Hook Parameters sheet ──────────────────────────────────────────────────────

def _write_hook_parameters_sheet(ws: Worksheet) -> tuple[str, str]:
    """
    Returns (a_range, key_range) — the exact cell ranges the seed rows
    landed in, so _write_bbs_sheet's INDEX/MATCH formulas always point
    at the real data even if HOOK_PARAMETER_SEED grows or shrinks.
    """
    ws["A1"] = "Hook Parameters — editable reference table"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = (
        "Edit column C to change a hook length. The BBS sheet looks these "
        "up by Key (column D) and never hardcodes a hook length itself."
    )
    ws["A2"].font = NOTE_FONT

    headers = ["Shape", "Steel", "A (mm)", "Key"]
    header_row = 4
    for col, text in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=text)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER

    row = header_row + 1
    for shape, steel, value in HOOK_PARAMETER_SEED:
        ws.cell(row=row, column=1, value=shape)
        ws.cell(row=row, column=2, value=steel)
        a_cell = ws.cell(row=row, column=3, value=value)
        a_cell.font = INPUT_FONT
        a_cell.fill = INPUT_FILL
        ws.cell(row=row, column=4, value=f'=A{row}&"_"&B{row}')
        row += 1

    first_data_row = header_row + 1
    last_data_row  = row - 1
    ws["A" + str(row + 1)] = (
        f"To add a new bar size or shape: add a row above row {row}, "
        f"filling Shape/Steel/A (mm) — Key fills itself."
    )
    ws["A" + str(row + 1)].font = NOTE_FONT

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 14

    a_range   = f"'Hook Parameters'!$C${first_data_row}:$C${last_data_row}"
    key_range = f"'Hook Parameters'!$D${first_data_row}:$D${last_data_row}"
    return a_range, key_range


# ── BBS sheet ──────────────────────────────────────────────────────────────────

def _write_bbs_sheet(ws: Worksheet, beams: list[Beam], hook_range_a: str, hook_range_key: str) -> int:
    ws["A1"] = "BAR BENDING SCHEDULE TO BS8666:2005"
    ws["A1"].font = TITLE_FONT

    header_row = 3
    for col, text in enumerate(BBS_HEADERS, start=1):
        cell = ws.cell(row=header_row, column=col, value=text)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER

    row = header_row + 1
    for beam in beams:
        first_row_for_beam = True
        for bar in beam.bars:
            r = row
            ws.cell(row=r, column=1, value=beam.id if first_row_for_beam else None)
            first_row_for_beam = False

            ws.cell(row=r, column=2, value=bar.mark)
            ws.cell(row=r, column=3, value=f"T{bar.diameter}")
            ws.cell(row=r, column=4, value=1)          # No. of members — genuinely 1 for every beam on this drawing
            ws.cell(row=r, column=5, value=bar.quantity)
            ws.cell(row=r, column=6, value=f"=D{r}*E{r}")
            ws.cell(row=r, column=7, value=f'=IF((I{r}+J{r}+K{r}+L{r}+M{r})<=12000,(I{r}+J{r}+K{r}+L{r}+M{r}),">12000")')
            ws.cell(row=r, column=8, value=bar.shape)

            _write_dimension(ws, r, 9,  bar.dim_a_mm, bar.dim_a_lookup_key, hook_range_a, hook_range_key)
            _write_dimension(ws, r, 10, bar.dim_b_mm, None, hook_range_a, hook_range_key)
            _write_dimension(ws, r, 11, bar.dim_c_mm, bar.dim_c_lookup_key, hook_range_a, hook_range_key)
            _write_dimension(ws, r, 12, bar.dim_d_mm, None, hook_range_a, hook_range_key)
            # column M ("E") intentionally left blank — unused by any
            # shape currently in scope (00/11/21/51); reserved for a
            # future shape that needs a 5th dimension.

            ws.cell(row=r, column=14, value=f"=G{r}*F{r}")

            for col in range(6, 15):
                ws.cell(row=r, column=col).alignment = Alignment(horizontal="center")

            row += 1

    last_row = row - 1
    for col in range(1, len(BBS_HEADERS) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 14
    ws.column_dimensions["A"].width = 12
    ws.freeze_panes = f"A{header_row + 1}"

    return last_row


def _write_dimension(
    ws:              Worksheet,
    row:             int,
    col:             int,
    value:           Optional[float],
    lookup_key:      Optional[str],
    hook_range_a:    str,
    hook_range_key:  str,
) -> None:
    """
    Writes one A/B/C/D cell: a plain number if length_calculator.py /
    shape_resolver.py already resolved it, or — for a hook dimension
    only — an INDEX/MATCH formula against the Hook Parameters sheet
    using the bar's own "{shape}_T{diameter}" key. Never both.
    """
    cell = ws.cell(row=row, column=col)
    if lookup_key is not None:
        cell.value = f'=IFERROR(INDEX({hook_range_a}, MATCH("{lookup_key}", {hook_range_key}, 0)), "")'
    elif value is not None:
        cell.value = round(value, 1)


# ── Steel Summary sheet ─────────────────────────────────────────────────────────

def _write_steel_summary_sheet(ws: Worksheet, beams: list[Beam], bbs_last_row: int) -> None:
    ws["A1"] = "STEEL SUMMARY"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = (
        "Total length is summed from the BBS sheet's own Total Length column; "
        f"unit weight uses the standard BS4449 relationship kg/m = d\u00b2/{STEEL_UNIT_WEIGHT_DIVISOR}."
    )
    ws["A2"].font = NOTE_FONT

    headers = ["Bar size", "Total length (m)", "Unit weight (kg/m)", "Total weight (kg)"]
    header_row = 4
    for col, text in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=text)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER

    diameters = sorted({bar.diameter for beam in beams for bar in beam.bars})

    row = header_row + 1
    for dia in diameters:
        size_label = f"T{dia}"
        ws.cell(row=row, column=1, value=size_label)
        ws.cell(row=row, column=2,
                value=f'=SUMIF(BBS!$C$4:$C${bbs_last_row},"{size_label}",BBS!$N$4:$N${bbs_last_row})/1000')
        ws.cell(row=row, column=3, value=f"=({dia}^2)/{STEEL_UNIT_WEIGHT_DIVISOR}")
        ws.cell(row=row, column=4, value=f"=B{row}*C{row}")
        row += 1

    total_row = row + 1
    ws.cell(row=total_row, column=1, value="TOTAL").font = HEADER_FONT
    ws.cell(row=total_row, column=4, value=f"=SUM(D{header_row + 1}:D{row - 1})").font = HEADER_FONT

    for col in range(1, 5):
        ws.column_dimensions[get_column_letter(col)].width = 18